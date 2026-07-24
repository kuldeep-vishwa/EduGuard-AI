"""
blueprints/faculty/routes.py – EduGuard AI Faculty Blueprint
=============================================================
Faculty dashboard, attendance, assignments, feedback,
question bank and assessment scheduling.
"""

from datetime import datetime, date
from functools import wraps
from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, jsonify)
from flask_login import login_required, current_user
from database import db
from models import (Faculty, Student, Subject, Unit, Department,
                    Attendance, Assignment, FacultyFeedback, Question,
                    Assessment, AssessmentAttempt, AssessmentAnswer,
                    FacultySubject, AttendanceStatusEnum, RoleEnum)
from forms import (AttendanceForm, AssignmentForm, FacultyFeedbackForm,
                   QuestionForm, AssessmentForm)

faculty_bp = Blueprint("faculty", __name__, url_prefix="/faculty")


# ── Access guard ──────────────────────────────────────────────────────────────

def faculty_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_faculty:
            flash("Faculty access required.", "danger")
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated


def get_current_faculty():
    """Helper: return Faculty record for logged-in user."""
    return Faculty.query.filter_by(user_id=current_user.id).first()


# ── Dashboard ─────────────────────────────────────────────────────────────────

@faculty_bp.route("/dashboard")
@login_required
@faculty_required
def dashboard():
    faculty = get_current_faculty()
    if not faculty:
        flash("Faculty profile not found.", "danger")
        return redirect(url_for("auth.logout"))

    from sqlalchemy import func

    # Subjects taught by this faculty
    subjects = [fs.subject for fs in faculty.subjects_taught.all()]
    subject_ids = [s.id for s in subjects]

    # Students in those subjects (via attendance)
    student_ids_rows = (db.session.query(Attendance.student_id)
                        .filter(Attendance.faculty_id == faculty.id)
                        .distinct().all())
    student_ids    = [s[0] for s in student_ids_rows]
    total_students = len(student_ids)

    # Upcoming assessments
    upcoming = (Assessment.query
                .filter_by(faculty_id=faculty.id, is_active=True)
                .filter(Assessment.scheduled_date >= date.today())
                .order_by(Assessment.scheduled_date)
                .limit(5).all())

    # Question bank count
    question_count = Question.query.filter_by(faculty_id=faculty.id, is_active=True).count()

    # Recent assessment results
    recent_attempts = (AssessmentAttempt.query
                       .join(Assessment)
                       .filter(Assessment.faculty_id == faculty.id,
                               AssessmentAttempt.is_completed == True)
                       .order_by(AssessmentAttempt.submitted_at.desc())
                       .limit(8).all())

    # ── AI Risk summary for this faculty's students ───────────────────────────
    risk_high   = Student.query.filter(Student.id.in_(student_ids), Student.risk_level == "high").count()
    risk_medium = Student.query.filter(Student.id.in_(student_ids), Student.risk_level == "medium").count()
    risk_low    = Student.query.filter(Student.id.in_(student_ids), Student.risk_level == "low").count()

    # ── Attendance trend: last 7 subjects (average present %) ────────────────
    att_trend_labels = []
    att_trend_data   = []
    for subj in subjects[:7]:
        total = db.session.query(func.count(Attendance.id)).filter_by(
            faculty_id=faculty.id, subject_id=subj.id).scalar() or 0
        present = db.session.query(func.count(Attendance.id)).filter_by(
            faculty_id=faculty.id, subject_id=subj.id, status=AttendanceStatusEnum.PRESENT).scalar() or 0
        pct = round((present / total * 100), 1) if total else 0
        att_trend_labels.append(subj.code)
        att_trend_data.append(pct)

    # ── Assessment score trend across assessments ─────────────────────────────
    assess_labels = []
    assess_data   = []
    for attempt in recent_attempts[:8]:
        assess_labels.append(attempt.assessment.title[:12])
        assess_data.append(attempt.percentage)

    return render_template("faculty/dashboard.html",
                           title="Faculty Dashboard",
                           faculty=faculty,
                           subjects=subjects,
                           total_students=total_students,
                           upcoming=upcoming,
                           question_count=question_count,
                           recent_attempts=recent_attempts,
                           risk_high=risk_high,
                           risk_medium=risk_medium,
                           risk_low=risk_low,
                           att_trend_labels=att_trend_labels,
                           att_trend_data=att_trend_data,
                           assess_labels=assess_labels,
                           assess_data=assess_data)


# ═════════════════════════════════════════════════════════════════════════════
# ATTENDANCE
# ═════════════════════════════════════════════════════════════════════════════

@faculty_bp.route("/attendance", methods=["GET", "POST"])
@login_required
@faculty_required
def attendance():
    faculty = get_current_faculty()
    form = AttendanceForm()
    subject_choices = [
        (fs.subject_id, fs.subject.name) for fs in faculty.subjects_taught.all()
    ]
    # WTForms SelectField requires at least one choice; provide a placeholder when empty
    form.subject_id.choices = subject_choices or [(0, "— No subjects assigned —")]

    if not subject_choices:
        flash("No subjects are assigned to you yet. Ask admin to assign subjects.", "warning")
        return render_template("faculty/attendance.html",
                               form=form, faculty=faculty, title="Attendance")

    if form.validate_on_submit():
        selected_subject = Subject.query.get(form.subject_id.data)
        selected_date = form.date.data
        students = Student.query.filter_by(is_active=True).all()
        return render_template("faculty/mark_attendance.html",
                               form=form,
                               faculty=faculty,
                               students=students,
                               subject=selected_subject,
                               att_date=selected_date,
                               title="Mark Attendance")

    return render_template("faculty/attendance.html",
                           form=form, faculty=faculty,
                           title="Attendance")


@faculty_bp.route("/attendance/save", methods=["POST"])
@login_required
@faculty_required
def save_attendance():
    faculty = get_current_faculty()
    subject_id = request.form.get("subject_id", type=int)
    att_date_str = request.form.get("att_date")
    att_date = datetime.strptime(att_date_str, "%Y-%m-%d").date()

    students = Student.query.filter_by(is_active=True).all()
    saved = 0
    for student in students:
        status = request.form.get(f"status_{student.id}", AttendanceStatusEnum.ABSENT)
        remarks = request.form.get(f"remarks_{student.id}", "")
        # Upsert attendance
        existing = Attendance.query.filter_by(
            student_id=student.id, subject_id=subject_id, date=att_date
        ).first()
        if existing:
            existing.status = status
            existing.remarks = remarks
        else:
            record = Attendance(
                student_id=student.id,
                subject_id=subject_id,
                faculty_id=faculty.id,
                date=att_date,
                status=status,
                remarks=remarks
            )
            db.session.add(record)
        saved += 1

    db.session.commit()
    flash(f"Attendance saved for {saved} students.", "success")
    return redirect(url_for("faculty.attendance_history"))


@faculty_bp.route("/attendance/history")
@login_required
@faculty_required
def attendance_history():
    faculty = get_current_faculty()
    subject_id = request.args.get("subject_id", type=int)
    att_date_str = request.args.get("date")

    query = Attendance.query.filter_by(faculty_id=faculty.id)
    if subject_id:
        query = query.filter_by(subject_id=subject_id)
    if att_date_str:
        try:
            att_date = datetime.strptime(att_date_str, "%Y-%m-%d").date()
            query = query.filter_by(date=att_date)
        except ValueError:
            pass

    records = query.order_by(Attendance.date.desc()).limit(200).all()
    subjects = [fs.subject for fs in faculty.subjects_taught.all()]

    return render_template("faculty/attendance_history.html",
                           records=records,
                           subjects=subjects,
                           faculty=faculty,
                           title="Attendance History")


# ═════════════════════════════════════════════════════════════════════════════
# ASSIGNMENTS
# ═════════════════════════════════════════════════════════════════════════════

@faculty_bp.route("/assignments", methods=["GET", "POST"])
@login_required
@faculty_required
def assignments():
    faculty = get_current_faculty()
    form = AssignmentForm()
    student_list = Student.query.filter_by(is_active=True).order_by(Student.full_name).all()
    subject_choices = [(fs.subject_id, fs.subject.name) for fs in faculty.subjects_taught.all()]
    form.student_id.choices = [(s.id, f"{s.student_id} – {s.full_name}") for s in student_list] \
                              or [(0, "— No students —")]
    form.subject_id.choices = subject_choices or [(0, "— No subjects assigned —")]

    if form.validate_on_submit():
        assignment = Assignment(
            student_id=form.student_id.data,
            subject_id=form.subject_id.data,
            faculty_id=faculty.id,
            title=form.title.data,
            max_marks=form.max_marks.data,
            obtained_marks=form.obtained_marks.data,
            submission_date=form.submission_date.data,
            remarks=form.remarks.data
        )
        db.session.add(assignment)
        db.session.commit()
        flash("Assignment marks saved.", "success")
        return redirect(url_for("faculty.assignments"))

    recent = (Assignment.query
              .filter_by(faculty_id=faculty.id)
              .order_by(Assignment.created_at.desc())
              .limit(20).all())
    return render_template("faculty/assignments.html",
                           form=form, recent=recent, faculty=faculty,
                           title="Assignment Marks")


# ═════════════════════════════════════════════════════════════════════════════
# FACULTY FEEDBACK
# ═════════════════════════════════════════════════════════════════════════════

@faculty_bp.route("/feedback", methods=["GET", "POST"])
@login_required
@faculty_required
def feedback():
    faculty = get_current_faculty()
    form = FacultyFeedbackForm()
    student_list = Student.query.filter_by(is_active=True).order_by(Student.full_name).all()
    form.student_id.choices = [(s.id, f"{s.student_id} – {s.full_name}") for s in student_list] \
                              or [(0, "— No students —")]
    form.subject_id.choices = [(0, "General")] + [
        (fs.subject_id, fs.subject.name) for fs in faculty.subjects_taught.all()
    ]

    if form.validate_on_submit():
        fb = FacultyFeedback(
            student_id=form.student_id.data,
            faculty_id=faculty.id,
            subject_id=form.subject_id.data or None,
            class_participation=form.class_participation.data,
            subject_understanding=form.subject_understanding.data,
            assignment_quality=form.assignment_quality.data,
            learning_progress=form.learning_progress.data,
            comments=form.comments.data,
            feedback_date=form.feedback_date.data or date.today()
        )
        fb.compute_overall()
        db.session.add(fb)
        db.session.commit()
        flash("Feedback submitted.", "success")
        return redirect(url_for("faculty.feedback"))

    recent_feedbacks = (FacultyFeedback.query
                        .filter_by(faculty_id=faculty.id)
                        .order_by(FacultyFeedback.created_at.desc())
                        .limit(15).all())
    return render_template("faculty/feedback.html",
                           form=form, recent_feedbacks=recent_feedbacks,
                           faculty=faculty, title="Faculty Feedback")


# ═════════════════════════════════════════════════════════════════════════════
# QUESTION BANK
# ═════════════════════════════════════════════════════════════════════════════

@faculty_bp.route("/questions")
@login_required
@faculty_required
def questions():
    faculty = get_current_faculty()
    subject_id = request.args.get("subject_id", type=int)
    unit_id = request.args.get("unit_id", type=int)
    difficulty = request.args.get("difficulty", "")
    page = request.args.get("page", 1, type=int)

    query = Question.query.filter_by(faculty_id=faculty.id)
    if subject_id:
        query = query.filter_by(subject_id=subject_id)
    if unit_id:
        query = query.filter_by(unit_id=unit_id)
    if difficulty:
        query = query.filter_by(difficulty=difficulty)

    questions_page = query.order_by(Question.created_at.desc()).paginate(page=page, per_page=15)
    subjects = [fs.subject for fs in faculty.subjects_taught.all()]

    return render_template("faculty/questions.html",
                           questions=questions_page,
                           subjects=subjects,
                           faculty=faculty,
                           filters={"subject_id": subject_id, "unit_id": unit_id, "difficulty": difficulty},
                           title="Question Bank")


@faculty_bp.route("/questions/add", methods=["GET", "POST"])
@login_required
@faculty_required
def add_question():
    faculty = get_current_faculty()
    form = QuestionForm()
    subject_choices = [(fs.subject_id, fs.subject.name)
                       for fs in faculty.subjects_taught.all()]
    form.subject_id.choices = subject_choices or [(0, "— No subjects assigned —")]

    if not subject_choices:
        flash("No subjects are assigned to you yet. Ask admin to assign subjects.", "warning")
        form.unit_id.choices = [(0, "— No units —")]
        return render_template("faculty/question_form.html",
                               form=form, faculty=faculty, title="Add Question")

    # Pre-populate units if subject_id provided via GET param, else use first subject
    subject_id = request.args.get("subject_id", type=int)
    if subject_id and any(c[0] == subject_id for c in subject_choices):
        units = Unit.query.filter_by(subject_id=subject_id).order_by(Unit.unit_number).all()
        form.unit_id.choices = [(u.id, f"Unit {u.unit_number}: {u.name}") for u in units]
        form.subject_id.data = subject_id
    else:
        units = Unit.query.filter_by(subject_id=subject_choices[0][0]).order_by(Unit.unit_number).all()
        form.unit_id.choices = [(u.id, f"Unit {u.unit_number}: {u.name}") for u in units] \
                               or [(0, "— No units —")]

    if form.validate_on_submit():
        question = Question(
            subject_id=form.subject_id.data,
            unit_id=form.unit_id.data,
            faculty_id=faculty.id,
            question_text=form.question_text.data,
            question_type=form.question_type.data,
            difficulty=form.difficulty.data,
            option_a=form.option_a.data,
            option_b=form.option_b.data,
            option_c=form.option_c.data,
            option_d=form.option_d.data,
            correct_answer=form.correct_answer.data,
            explanation=form.explanation.data,
            is_active=form.is_active.data
        )
        db.session.add(question)
        db.session.commit()
        flash("Question added to bank.", "success")
        return redirect(url_for("faculty.questions"))

    return render_template("faculty/question_form.html",
                           form=form, faculty=faculty, title="Add Question")


@faculty_bp.route("/questions/<int:q_id>/edit", methods=["GET", "POST"])
@login_required
@faculty_required
def edit_question(q_id):
    faculty = get_current_faculty()
    question = Question.query.filter_by(id=q_id, faculty_id=faculty.id).first_or_404()
    form = QuestionForm(obj=question)
    form.subject_id.choices = [(fs.subject_id, fs.subject.name)
                               for fs in faculty.subjects_taught.all()] \
                              or [(question.subject_id, question.subject.name)]
    units = Unit.query.filter_by(subject_id=question.subject_id).order_by(Unit.unit_number).all()
    form.unit_id.choices = [(u.id, f"Unit {u.unit_number}: {u.name}") for u in units] \
                           or [(question.unit_id, question.unit.name)]

    if form.validate_on_submit():
        question.question_text = form.question_text.data
        question.question_type = form.question_type.data
        question.difficulty = form.difficulty.data
        question.subject_id = form.subject_id.data
        question.unit_id = form.unit_id.data
        question.option_a = form.option_a.data
        question.option_b = form.option_b.data
        question.option_c = form.option_c.data
        question.option_d = form.option_d.data
        question.correct_answer = form.correct_answer.data
        question.explanation = form.explanation.data
        question.is_active = form.is_active.data
        db.session.commit()
        flash("Question updated.", "success")
        return redirect(url_for("faculty.questions"))

    return render_template("faculty/question_form.html",
                           form=form, faculty=faculty, edit=True,
                           question=question, title="Edit Question")


@faculty_bp.route("/questions/<int:q_id>/delete", methods=["POST"])
@login_required
@faculty_required
def delete_question(q_id):
    faculty = get_current_faculty()
    question = Question.query.filter_by(id=q_id, faculty_id=faculty.id).first_or_404()
    question.is_active = False
    db.session.commit()
    flash("Question removed.", "info")
    return redirect(url_for("faculty.questions"))


# ═════════════════════════════════════════════════════════════════════════════
# ASSESSMENT SCHEDULER
# ═════════════════════════════════════════════════════════════════════════════

@faculty_bp.route("/assessments")
@login_required
@faculty_required
def assessments():
    faculty = get_current_faculty()
    page = request.args.get("page", 1, type=int)
    assessments_page = (Assessment.query
                        .filter_by(faculty_id=faculty.id)
                        .order_by(Assessment.scheduled_date.desc())
                        .paginate(page=page, per_page=15))
    return render_template("faculty/assessments.html",
                           assessments=assessments_page,
                           faculty=faculty,
                           title="Assessments")


@faculty_bp.route("/assessments/add", methods=["GET", "POST"])
@login_required
@faculty_required
def add_assessment():
    faculty = get_current_faculty()
    form = AssessmentForm()
    subject_choices = [(fs.subject_id, fs.subject.name)
                       for fs in faculty.subjects_taught.all()]
    form.subject_id.choices = subject_choices or [(0, "— No subjects assigned —")]
    form.unit_id.choices = [(0, "All Units")]

    if not subject_choices:
        flash("No subjects are assigned to you yet. Ask admin to assign subjects.", "warning")
        return render_template("faculty/assessment_form.html",
                               form=form, faculty=faculty, title="Schedule Assessment")

    units = Unit.query.filter_by(subject_id=subject_choices[0][0]).order_by(Unit.unit_number).all()
    form.unit_id.choices += [(u.id, f"Unit {u.unit_number}: {u.name}") for u in units]

    if form.validate_on_submit():
        assessment = Assessment(
            title=form.title.data,
            subject_id=form.subject_id.data,
            unit_id=form.unit_id.data or None,
            faculty_id=faculty.id,
            scheduled_date=form.scheduled_date.data,
            start_time=form.start_time.data,
            duration_minutes=form.duration_minutes.data,
            total_questions=form.total_questions.data,
            easy_count=form.easy_count.data,
            medium_count=form.medium_count.data,
            hard_count=form.hard_count.data,
            max_marks=form.max_marks.data,
            is_active=form.is_active.data
        )
        db.session.add(assessment)
        db.session.commit()
        flash("Assessment scheduled successfully.", "success")
        return redirect(url_for("faculty.assessments"))

    return render_template("faculty/assessment_form.html",
                           form=form, faculty=faculty, title="Schedule Assessment")


@faculty_bp.route("/assessments/<int:a_id>/edit", methods=["GET", "POST"])
@login_required
@faculty_required
def edit_assessment(a_id):
    faculty = get_current_faculty()
    assessment = Assessment.query.filter_by(id=a_id, faculty_id=faculty.id).first_or_404()
    form = AssessmentForm(obj=assessment)
    form.subject_id.choices = [(fs.subject_id, fs.subject.name)
                               for fs in faculty.subjects_taught.all()] \
                              or [(assessment.subject_id, assessment.subject.name)]
    units = Unit.query.filter_by(subject_id=assessment.subject_id).order_by(Unit.unit_number).all()
    form.unit_id.choices = [(0, "All Units")] + [(u.id, f"Unit {u.unit_number}: {u.name}") for u in units]

    if form.validate_on_submit():
        assessment.title = form.title.data
        assessment.subject_id = form.subject_id.data
        assessment.unit_id = form.unit_id.data or None
        assessment.scheduled_date = form.scheduled_date.data
        assessment.start_time = form.start_time.data
        assessment.duration_minutes = form.duration_minutes.data
        assessment.total_questions = form.total_questions.data
        assessment.easy_count = form.easy_count.data
        assessment.medium_count = form.medium_count.data
        assessment.hard_count = form.hard_count.data
        assessment.max_marks = form.max_marks.data
        assessment.is_active = form.is_active.data
        db.session.commit()
        flash("Assessment updated.", "success")
        return redirect(url_for("faculty.assessments"))

    return render_template("faculty/assessment_form.html",
                           form=form, faculty=faculty, edit=True,
                           assessment=assessment, title="Edit Assessment")


@faculty_bp.route("/assessments/<int:a_id>/results")
@login_required
@faculty_required
def assessment_results(a_id):
    faculty = get_current_faculty()
    assessment = Assessment.query.filter_by(id=a_id, faculty_id=faculty.id).first_or_404()
    attempts = (assessment.attempts
                .filter_by(is_completed=True)
                .order_by(AssessmentAttempt.score.desc()).all())
    avg_score = sum(a.percentage for a in attempts) / len(attempts) if attempts else 0
    return render_template("faculty/assessment_results.html",
                           assessment=assessment,
                           attempts=attempts,
                           avg_score=round(avg_score, 2),
                           faculty=faculty,
                           title="Assessment Results")


# ═════════════════════════════════════════════════════════════════════════════
# STUDENT RECORDS  (faculty view)
# ═════════════════════════════════════════════════════════════════════════════

@faculty_bp.route("/students")
@login_required
@faculty_required
def student_list():
    faculty = get_current_faculty()
    students = Student.query.filter_by(is_active=True).order_by(Student.full_name).all()
    return render_template("faculty/students.html",
                           students=students, faculty=faculty,
                           title="Students")


@faculty_bp.route("/students/<int:student_id>")
@login_required
@faculty_required
def view_student(student_id):
    faculty = get_current_faculty()
    student = Student.query.get_or_404(student_id)
    attendance = student.attendance_records.order_by(Attendance.date.desc()).limit(20).all()
    assignments = student.assignments.order_by(Assignment.created_at.desc()).limit(10).all()
    feedbacks = (FacultyFeedback.query
                 .filter_by(student_id=student.id, faculty_id=faculty.id)
                 .order_by(FacultyFeedback.created_at.desc()).all())
    attempts = (student.assessment_attempts
                .filter_by(is_completed=True)
                .order_by(AssessmentAttempt.submitted_at.desc()).limit(10).all())
    return render_template("faculty/student_detail.html",
                           student=student,
                           faculty=faculty,
                           attendance=attendance,
                           assignments=assignments,
                           feedbacks=feedbacks,
                           attempts=attempts,
                           title=f"Student – {student.full_name}")


# ── AJAX: units by subject ─────────────────────────────────────────────────

@faculty_bp.route("/api/units/<int:subject_id>")
@login_required
def get_units_for_subject(subject_id):
    units = Unit.query.filter_by(subject_id=subject_id).order_by(Unit.unit_number).all()
    return jsonify([{"id": u.id, "name": f"Unit {u.unit_number}: {u.name}"} for u in units])


# ═════════════════════════════════════════════════════════════════════════════
# AI QUESTION GENERATOR
# ═════════════════════════════════════════════════════════════════════════════

@faculty_bp.route("/questions/generate-ai", methods=["GET", "POST"])
@login_required
@faculty_required
def generate_questions_ai():
    """
    AI Question Generator – uses IBM Granite to generate quiz questions.
    GET  → show form
    POST → call ai_service.generate_ai_questions, display results for review
    POST with action=save_selected → save checked questions to the Question Bank
    """
    faculty = get_current_faculty()
    subjects = [fs.subject for fs in faculty.subjects_taught.all()]

    if not subjects:
        flash("No subjects are assigned to you. Ask admin to assign subjects.", "warning")
        return redirect(url_for("faculty.questions"))

    # ── Save selected AI questions to the bank ───────────────────────────────
    if request.method == "POST" and request.form.get("action") == "save_selected":
        saved = 0
        subject_id = request.form.get("subject_id", type=int)
        unit_id    = request.form.get("unit_id",    type=int)
        import json as _json

        selected_indices = request.form.getlist("selected_q")
        questions_json   = request.form.get("questions_json", "[]")
        try:
            all_questions = _json.loads(questions_json)
        except Exception:
            all_questions = []

        for idx_str in selected_indices:
            try:
                idx = int(idx_str)
                q   = all_questions[idx]
            except (ValueError, IndexError, TypeError):
                continue
            if not q.get("question_text"):
                continue
            question = Question(
                subject_id=subject_id,
                unit_id=unit_id,
                faculty_id=faculty.id,
                question_text=q["question_text"],
                question_type=q.get("question_type", "mcq"),
                difficulty=q.get("difficulty", "medium"),
                option_a=q.get("option_a", ""),
                option_b=q.get("option_b", ""),
                option_c=q.get("option_c", ""),
                option_d=q.get("option_d", ""),
                correct_answer=q.get("correct_answer", ""),
                explanation=q.get("explanation", ""),
                is_active=True
            )
            db.session.add(question)
            saved += 1

        db.session.commit()
        flash(f"{saved} AI-generated question(s) added to your Question Bank.", "success")
        return redirect(url_for("faculty.questions"))

    # ── Generate questions via IBM Granite ────────────────────────────────────
    ai_questions = []
    gen_params   = {}

    if request.method == "POST":
        subject_id    = request.form.get("subject_id", type=int)
        unit_id       = request.form.get("unit_id",    type=int)
        difficulty    = request.form.get("difficulty", "medium")
        question_type = request.form.get("question_type", "mcq")
        count         = request.form.get("count", 5, type=int)
        topic_hint    = request.form.get("topic_hint", "").strip()

        subject = Subject.query.get(subject_id) if subject_id else None
        unit    = Unit.query.get(unit_id)    if unit_id    else None

        if not subject or not unit:
            flash("Please select a valid subject and unit.", "warning")
        else:
            gen_params = {
                "subject_id": subject_id,
                "unit_id": unit_id,
                "difficulty": difficulty,
                "question_type": question_type,
                "count": count,
                "topic_hint": topic_hint,
            }
            from ai_service import generate_ai_questions
            ai_questions = generate_ai_questions(
                subject_name=subject.name,
                unit_name=unit.name,
                difficulty=difficulty,
                question_type=question_type,
                count=count,
                topic_hint=topic_hint,
            )
            if ai_questions:
                flash(f"IBM Granite generated {len(ai_questions)} question(s). Review and save the ones you want.", "success")
            else:
                flash("IBM Granite is not available right now. Check your .env credentials and try again.", "warning")

    # Build subject/unit choices for the form
    subject_choices = [(s.id, f"{s.code} – {s.name}") for s in subjects]
    default_subject_id = int(request.form.get("subject_id", 0)) or subjects[0].id
    units = Unit.query.filter_by(subject_id=default_subject_id).order_by(Unit.unit_number).all()

    import json as _json
    return render_template(
        "faculty/generate_questions_ai.html",
        faculty=faculty,
        subjects=subjects,
        subject_choices=subject_choices,
        units=units,
        ai_questions=ai_questions,
        gen_params=gen_params,
        questions_json=_json.dumps(ai_questions),
        title="AI Question Generator"
    )


# ═════════════════════════════════════════════════════════════════════════════
# BULK IMPORT (CSV / XLSX) – Students
# ═════════════════════════════════════════════════════════════════════════════

@faculty_bp.route("/export/class-excel")
@login_required
@faculty_required
def export_class_excel():
    """
    Export the faculty's class list as a professionally formatted Excel file.
    Includes student info, attendance %, assessment avg, AI risk level,
    and faculty feedback – with risk colour-coding.
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("openpyxl is not installed. Run: pip install openpyxl>=3.1.0", "danger")
        return redirect(url_for("faculty.student_list"))

    from flask import current_app, send_file
    from models import AttendanceStatusEnum, AssessmentAttempt, FacultyFeedback
    from sqlalchemy import func
    from datetime import date

    faculty = get_current_faculty()
    students = Student.query.filter_by(is_active=True).order_by(Student.full_name).all()

    # ── Create workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Class Report"

    # Colour palette
    HEADER_BG   = "1F3864"   # dark navy
    HEADER_FG   = "FFFFFF"
    LOW_BG      = "C6EFCE"   # green
    MEDIUM_BG   = "FFEB9C"   # yellow
    HIGH_BG     = "FFC7CE"   # red
    ALT_BG      = "F2F2F2"   # light grey alternate row
    TITLE_BG    = "2E75B6"   # blue title bar

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell_style(cell, bg=None, fg="000000", bold=False, center=False, size=11):
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.border = border
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"EduGuard AI – Class Report | {faculty.full_name} | {date.today().strftime('%d %b %Y')}"
    cell_style(title_cell, bg=TITLE_BG, fg=HEADER_FG, bold=True, center=True, size=13)
    ws.row_dimensions[1].height = 28

    # ── Sub-header ───────────────────────────────────────────────────────────
    ws.merge_cells("A2:K2")
    dept_cell = ws["A2"]
    dept_cell.value = (
        f"Department: {faculty.department.name if faculty.department else 'N/A'} | "
        f"Subjects: {', '.join(s.name for s in [fs.subject for fs in faculty.subjects_taught.all()])}"
    )
    cell_style(dept_cell, bg="D9E2F3", center=True)
    ws.row_dimensions[2].height = 18

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        "#", "Student Name", "Student ID", "Department", "Semester",
        "Attendance %", "Avg Score %", "Assignments", "Faculty Feedback",
        "Health Score", "Risk Level"
    ]
    ws.row_dimensions[3].height = 22
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell_style(cell, bg=HEADER_BG, fg=HEADER_FG, bold=True, center=True)

    # Column widths
    widths = [5, 22, 14, 18, 10, 14, 13, 13, 17, 13, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_num, student in enumerate(students, 1):
        r = row_num + 3

        # Attendance
        total_att = student.attendance_records.count()
        present   = student.attendance_records.filter_by(status=AttendanceStatusEnum.PRESENT).count()
        att_pct   = round((present / total_att * 100), 1) if total_att else 0.0

        # Assessment average
        avg_score = db.session.query(func.avg(AssessmentAttempt.percentage)).filter_by(
            student_id=student.id, is_completed=True
        ).scalar() or 0
        avg_score = round(avg_score, 1)

        # Assignment average
        assignments = student.assignments.filter(Assignment.obtained_marks.isnot(None)).all()
        if assignments:
            asgn_avg = round(
                sum(a.obtained_marks / a.max_marks * 100 for a in assignments if a.max_marks) / len(assignments), 1
            )
        else:
            asgn_avg = 0.0

        # Faculty feedback (this faculty only)
        fb_avg = db.session.query(func.avg(FacultyFeedback.overall_rating)).filter_by(
            student_id=student.id, faculty_id=faculty.id
        ).scalar() or 0
        fb_avg = round(fb_avg, 2)

        health_score = student.academic_health_score or 0
        risk_level   = student.risk_level or "unknown"

        # Row colour based on risk
        if risk_level == "high":
            row_bg = HIGH_BG
        elif risk_level == "medium":
            row_bg = MEDIUM_BG
        elif risk_level == "low":
            row_bg = LOW_BG
        else:
            row_bg = ALT_BG if row_num % 2 == 0 else None

        dept_name = student.department.name if student.department else "—"

        row_data = [
            row_num,
            student.full_name,
            student.student_id,
            dept_name,
            student.current_semester,
            f"{att_pct}%",
            f"{avg_score}%",
            f"{asgn_avg}%",
            f"{fb_avg}/5",
            f"{health_score:.1f}/100",
            risk_level.upper() if risk_level != "unknown" else "N/A",
        ]

        ws.row_dimensions[r].height = 18
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col_num, value=value)
            cell_style(cell, bg=row_bg, center=(col_num != 2))

    # ── Summary row ────────────────────────────────────────────────────────────
    summary_row = len(students) + 4
    ws.merge_cells(f"A{summary_row}:E{summary_row}")
    summary_cell = ws[f"A{summary_row}"]
    summary_cell.value = f"Total Students: {len(students)}"
    cell_style(summary_cell, bg="D9E2F3", bold=True)

    # ── Legend ─────────────────────────────────────────────────────────────────
    legend_row = summary_row + 2
    ws.cell(row=legend_row, column=1, value="Risk Legend:").font = Font(bold=True)
    for col, (label, bg) in enumerate(
        [("Low Risk", LOW_BG), ("Medium Risk", MEDIUM_BG), ("High Risk", HIGH_BG)], 2
    ):
        cell = ws.cell(row=legend_row, column=col, value=label)
        cell_style(cell, bg=bg, center=True, bold=True)

    # ── Save to buffer ─────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"class_report_{faculty.faculty_id}_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
